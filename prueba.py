import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from ttkthemes import ThemedTk
import os
import pandas as pd

# Crear la ventana principal con el tema 'vistaero'
root = ThemedTk(theme="vistaero")

# Establecer el título de la ventana
root.title("Inflación")

# Obtener las dimensiones de la pantalla
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calcular el tamaño de la ventana como el 40% de la pantalla
window_width = int(screen_width * 0.7)
window_height = int(screen_height * 0.7)

# Centrar la ventana en la pantalla
position_top = int(screen_height / 2 - window_height / 2)
position_right = int(screen_width / 2 - window_width / 2)

# Establecer la geometría de la ventana
root.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")

# Crear un Frame para contener el botón, la mini previsualización, el Entry y el botón de aumento
frame = ttk.Frame(root)
frame.pack(expand=True, fill="both")

# Variables para almacenar el porcentaje
percentage_var = tk.DoubleVar()
percentage_var.set(0)  # Valor inicial

selected_filename = ""  # Variable para almacenar el nombre del archivo seleccionado


# Función para abrir el cuadro de diálogo de selección de archivos
def open_file_dialog():
    global selected_filename
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if filename:
        selected_filename = filename
        # Obtener solo el nombre del archivo sin la ruta
        file_name = os.path.basename(filename)
        # Actualizar el texto del botón con el nombre del archivo
        button.config(text=f"Seleccionado: {file_name}")

        # Mostrar mini previsualización
        show_preview(frame, filename)

        # Añadir Entry para ingresar el porcentaje
        add_percentage_entry(frame)

        # Añadir botón para aumentar el porcentaje
        add_increase_button(frame)

    else:
        # Restaurar el texto original si no se selecciona ningún archivo
        button.config(text="Seleccionar tabla")


# Función para mostrar la mini previsualización
def show_preview(container, filename):
    try:
        # Leer el archivo Excel utilizando pandas
        df = pd.read_excel(filename, skiprows=[0])

        # Crear un widget Treeview para mostrar las primeras 6 filas de la tabla
        tree = ttk.Treeview(
            container,
            columns=list(df.columns),
            show="headings",
            height=min(7, len(df) + 1),  # Mostrar las primeras 6 filas
        )

        # Configurar encabezados de columnas
        for col in df.columns:
            tree.heading(col, text=col)

        # Insertar datos en el Treeview (solo las primeras 6 filas)
        for index, row in df.head(6).iterrows():
            tree.insert("", "end", values=list(row))

        # Añadir el Treeview al contenedor
        tree.pack(expand=True, fill="both")

    except Exception as e:
        print(f"Error al mostrar la previsualización: {e}")


# Función para agregar el Entry para el porcentaje
def add_percentage_entry(container):
    entry_label = ttk.Label(container, text="Porcentaje a aumentar:")
    entry_label.pack(pady=5)

    entry = ttk.Entry(container, textvariable=percentage_var)
    entry.pack(pady=5, ipadx=10, ipady=5)


# Función para agregar el botón de aumento
def add_increase_button(container):
    increase_button = ttk.Button(
        container, text="Aumentar", command=increase_percentage
    )
    increase_button.pack(pady=5, ipadx=10, ipady=5)


# Función para aumentar el porcentaje
def increase_percentage():
    try:
        # Obtener el porcentaje ingresado
        percentage = percentage_var.get()

        # Realizar la acción de aumento (aquí puedes implementar tu lógica)
        print(f"Aumentando porcentaje en {percentage}% en la columna 'PRECIO'")

        # Leer el archivo Excel utilizando pandas
        df = pd.read_excel(selected_filename, skiprows=[0])

        # Verificar si la columna "PRECIO" existe en el DataFrame
        if "PRECIO" in df.columns:
            # Aumentar el porcentaje solo en los valores numéricos de la columna "PRECIO"
            df["PRECIO"] = df["PRECIO"].apply(
                lambda x: x * (1 + percentage / 100) if pd.api.types.is_number(x) else x
            )

            # Guardar el DataFrame modificado en un nuevo archivo Excel
            df.to_excel("new_file.xlsx", index=False)

            print(
                "Se ha creado un nuevo archivo 'new_file.xlsx' con el porcentaje actualizado."
            )

            # Mostrar la previsualización actualizada
            show_updated_preview(frame, "new_file.xlsx")

        else:
            print("La columna 'PRECIO' no existe en el DataFrame.")

    except Exception as e:
        print(f"Error al aumentar el porcentaje en la columna 'PRECIO': {e}")


def show_updated_preview(container, filename):
    try:
        # Leer el archivo Excel utilizando pandas
        df = pd.read_excel(filename)

        # Crear un widget Treeview para mostrar las primeras 6 filas de la tabla
        tree = ttk.Treeview(
            container,
            columns=list(df.columns),
            show="headings",
            height=min(7, len(df) + 1),  # Mostrar las primeras 6 filas
        )

        # Configurar encabezados de columnas
        for col in df.columns:
            tree.heading(col, text=col)

        # Insertar datos en el Treeview (solo las primeras 6 filas)
        for index, row in df.head(6).iterrows():
            tree.insert("", "end", values=list(row))

        # Añadir el Treeview al contenedor
        tree.pack(expand=True, fill="both")

        # Añadir botón de descarga
        download_button = ttk.Button(
            container, text="Descargar", command=lambda: download_file("new_file.xlsx")
        )
        download_button.pack(pady=5, ipadx=10, ipady=5)

    except Exception as e:
        print(f"Error al mostrar la previsualización actualizada: {e}")


from openpyxl import load_workbook
from tkinter import filedialog
import os


# Función para descargar el archivo y mostrar un mensaje en la pantalla
def download_file(filename):
    try:
        # Abrir el cuadro de diálogo para seleccionar la ubicación de descarga
        download_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
        )

        if download_path:
            # Copiar el archivo a la ubicación de descarga
            os.replace(filename, download_path)
            message_label.config(
                text=f"Archivo descargado con éxito en: {download_path}",
                foreground="green",
            )

            # Carga tu archivo Excel
            wb = load_workbook(download_path)

            # Selecciona la hoja en la que deseas agregar el encabezado de grupo
            sheet = wb.active

            # Inserta una nueva fila en la parte superior de la hoja
            sheet.insert_rows(1)

            # Combina las celdas para el encabezado de grupo
            sheet.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=len(sheet[2])
            )

            # Agrega el encabezado de grupo
            sheet.cell(row=1, column=1).value = "Tu Encabezado de Grupo"

            # Guarda el archivo Excel
            wb.save(download_path)

        else:
            message_label.config(text="Descarga cancelada.", foreground="black")

    except Exception as e:
        message_label.config(
            text=f"Error al descargar el archivo: {e}", foreground="red"
        )


# Crear el botón con estilo personalizado en el Frame
button = ttk.Button(frame, text="Seleccionar tabla", command=open_file_dialog)
button.pack(pady=5, ipadx=10, ipady=5)


# Crear una etiqueta para mostrar mensajes
message_label = ttk.Label(frame, text="", foreground="black")
message_label.pack(pady=5)

# Ejecutar la aplicación
root.mainloop()
